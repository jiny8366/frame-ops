"""
북촌 판매일지 Excel(모델번호·컬러번호·금액·결제방법) → `sales_import`용 CSV 텍스트.

시트명 `0410` = 4월 10일 등 MMDD. 상단 5행 안의 날짜(엑셀 datetime)가 있으면 그 날을 영업일로 씁니다.
컬럼 A~E: 번호, 모델번호, 컬러번호, 금액, 결제방법. (F~ 는 무시)

시간형 모델(`01:01`, `01:13` …)은 **표시만 시각처럼 보일 뿐 상품코드 조각은 항상 텍스트 `HH:MM`** 로 다룹니다.
엑셀이 셀을 「시간」 서식으로 저장해 `datetime`·`time`·일 단위 소수(하루 비율)로 읽히더라도 동일 문자열로 환원합니다.
`CX2197` 등 비시간 모델은 문자열 그대로 이어 붙입니다. (DB에 동일 product_code 필요)
"""

from __future__ import annotations

import csv
import io
import math
import os
import re
import tempfile
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from lib.constants import business_date_to_timestamptz

_SHEET_MMDD = re.compile(r"^(\d{2})(\d{2})$")


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return False


def normalize_model_cell_to_text(value: Any) -> str:
    """
    모델번호 셀 → 상품코드용 텍스트.

    엑셀 시각 서식은 내부적으로 time/datetime/하루 비율(float)일 수 있으나,
    `01:01` 과 같이 **항상 두 자리 시·분 텍스트**로 맞춘다. (DB `01:01-C01` 과 일치)
    """
    if _is_blank(value):
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        x = float(value)
        if 0.0 <= x < 1.0:
            secs = int(round(x * 24 * 3600)) % (24 * 3600)
            h = secs // 3600
            mm = (secs % 3600) // 60
            return f"{h:02d}:{mm:02d}"
        if x > 1.0:
            epoch = datetime(1899, 12, 30)
            dt = epoch + timedelta(days=x)
            return dt.strftime("%H:%M")
    return str(value).strip()


def normalize_color_code(color: str) -> str:
    s = (color or "").strip().upper()
    if not s.startswith("C"):
        return s
    rest = s[1:]
    if rest.isdigit() and len(rest) == 1:
        return f"C{int(rest):02d}"
    return s


def model_and_color_to_product_code(model_raw: Any, color_raw: Any) -> str:
    m = normalize_model_cell_to_text(model_raw)
    if _is_blank(color_raw):
        cs = ""
    elif isinstance(color_raw, str):
        cs = color_raw.strip()
    else:
        cs = str(color_raw).strip()
    c = normalize_color_code(cs)
    if not m or not c:
        raise ValueError("모델번호·컬러번호 필요")
    tm = re.match(r"^(\d{1,2}):(\d{2})$", m)
    if tm:
        h, mm = int(tm.group(1)), int(tm.group(2))
        m = f"{h:02d}:{mm:02d}"
    return f"{m}-{c}"


def _biz_date_from_sheet_name(sheet_name: str) -> date:
    m = _SHEET_MMDD.match(sheet_name.strip())
    if not m:
        return date.today()
    month, day = int(m.group(1)), int(m.group(2))
    y = date.today().year
    try:
        return date(y, month, day)
    except ValueError:
        return date(2026, month, day)


def _split_payment(amount: int, pay_raw: Any) -> tuple[int, int]:
    p = (str(pay_raw) if pay_raw is not None else "").strip()
    if not p:
        raise ValueError("결제방법 비어 있음")
    if p == "현금" or (p.startswith("현금") and "카드" not in p):
        return amount, 0
    if "카드" in p or p in ("신용카드", "체크카드"):
        return 0, amount
    raise ValueError(f"결제방법 인식 불가: {p!r}")


def _is_header_row(a: Any, b: Any) -> bool:
    sa = str(a).strip() if not _is_blank(a) else ""
    sb = str(b).strip() if not _is_blank(b) else ""
    return sa == "번호" and ("모델" in sb or sb == "모델번호")


def iter_bukchon_sales_rows(
    path: str | Path,
    sheet_name: str,
    *,
    store_code: str = "BKC01",
    clerk_note: str = "북촌 판매일지(안목)",
) -> tuple[date, list[dict[str, str]]]:
    """
    반환: (영업일, sales_import용 CSV 행 dict 목록).
    각 판매는 1행 1전표(라인 1개).
    """
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet_name!r} / 가능: {wb.sheetnames[:30]}")
        ws = wb[sheet_name]

        biz_from_cell: date | None = None
        header_seen = False
        empty_run = 0
        out: list[dict[str, str]] = []
        row_ix = 0

        for row in ws.iter_rows(min_row=1, max_col=5, values_only=True):
            row_ix += 1
            cells = list(row) + [None] * 5
            a, b, c, d, e = cells[:5]

            if row_ix <= 5 and biz_from_cell is None:
                v = a
                if isinstance(v, datetime):
                    biz_from_cell = v.date()
                elif isinstance(v, date):
                    biz_from_cell = v

            if not header_seen:
                if _is_header_row(a, b):
                    header_seen = True
                continue

            ms = normalize_model_cell_to_text(b)
            if not ms:
                empty_run += 1
                if empty_run >= 45:
                    break
                continue
            empty_run = 0

            if _is_blank(c) or _is_blank(d):
                continue
            try:
                amt = int(round(float(d)))
            except (TypeError, ValueError):
                continue
            if amt <= 0:
                continue

            try:
                pcode = model_and_color_to_product_code(b, c)
            except ValueError:
                continue

            try:
                cash, card = _split_payment(amt, e)
            except ValueError:
                continue

            biz = biz_from_cell or _biz_date_from_sheet_name(sheet_name)
            sold_iso = business_date_to_timestamptz(biz)

            rk = f"bukchon-{sheet_name}-{row_ix}"
            idem = f"bukchon-xlsx-{sheet_name}-{row_ix}"
            out.append(
                {
                    "receipt_key": rk,
                    "store_code": store_code,
                    "sold_at": sold_iso,
                    "product_code": pcode,
                    "quantity": "1",
                    "unit_price": str(amt),
                    "cash_amount": str(cash),
                    "card_amount": str(card),
                    "discount_total": "0",
                    "clerk_note": clerk_note,
                    "idempotency_key": idem,
                }
            )

        if not header_seen:
            raise ValueError("표 헤더(번호·모델번호 행)를 찾지 못했습니다.")
        return (biz_from_cell or _biz_date_from_sheet_name(sheet_name)), out
    finally:
        wb.close()


def bukchon_sales_xlsx_to_csv_text(
    path: str | Path,
    sheet_name: str,
    *,
    store_code: str = "BKC01",
    clerk_note: str = "북촌 판매일지(안목)",
) -> tuple[date, str]:
    biz, rows = iter_bukchon_sales_rows(
        path, sheet_name, store_code=store_code, clerk_note=clerk_note
    )
    if not rows:
        raise ValueError(f"유효 판매 행이 없습니다. 시트={sheet_name!r}")
    buf = io.StringIO()
    fieldnames = list(rows[0].keys())
    w = csv.DictWriter(buf, fieldnames=fieldnames, lineterminator="\n")
    w.writeheader()
    w.writerows(rows)
    return biz, buf.getvalue()


def list_mmdd_sheet_names(path: str | Path) -> list[str]:
    """`0410` 형태 시트만 골라 정렬."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = [n for n in wb.sheetnames if _SHEET_MMDD.match(str(n).strip())]
        return sorted(names)
    finally:
        wb.close()


def list_mmdd_sheet_names_from_bytes(data: bytes) -> list[str]:
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(data)
        return list_mmdd_sheet_names(path)
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


def bukchon_sales_xlsx_bytes_to_csv_text(
    data: bytes,
    sheet_name: str,
    *,
    store_code: str = "BKC01",
    clerk_note: str = "북촌 판매일지(안목)",
) -> tuple[date, str]:
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(data)
        return bukchon_sales_xlsx_to_csv_text(path, sheet_name, store_code=store_code, clerk_note=clerk_note)
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


def bukchon_sales_xlsx_all_mmdd_to_csv_text(path: str | Path) -> str:
    """동일 파일의 모든 MMDD 시트를 한 CSV로 이어 붙입니다. 판매 행 없는 시트는 건너뜁니다."""
    sheets = list_mmdd_sheet_names(path)
    if not sheets:
        raise ValueError("MMDD 형태 시트가 없습니다.")
    chunks: list[str] = []
    header_done = False
    for sn in sheets:
        try:
            _biz, csv = bukchon_sales_xlsx_to_csv_text(path, sn)
        except ValueError as e:
            if "유효 판매 행이 없습니다" in str(e):
                continue
            raise
        lines = [ln for ln in csv.splitlines() if ln.strip()]
        if len(lines) <= 1:
            continue
        if not header_done:
            chunks.extend(lines)
            header_done = True
        else:
            chunks.extend(lines[1:])
    if not chunks:
        raise ValueError("MMDD 시트에 적재할 판매 행이 한 건도 없습니다.")
    return "\n".join(chunks) + "\n"


def bukchon_sales_xlsx_all_mmdd_bytes_to_csv_text(data: bytes) -> str:
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(data)
        return bukchon_sales_xlsx_all_mmdd_to_csv_text(path)
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass

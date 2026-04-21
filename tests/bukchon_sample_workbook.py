"""실제 「판매일지_북촌」 xlsx 레이아웃에 맞춘 테스트용 워크북 생성."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook

# 사용자 제공 0410 시트와 동일한 4건(모델·컬러·금액·결제)
GOLDEN_0410_SALES: tuple[tuple[str, str, int, str], ...] = (
    (" 01:13", "C2", 50000, "현금"),
    ("04:20", "C2", 50000, "현금"),
    ("07:04", "C2", 50000, "카드"),
    ("04:35", "C2", 50000, "카드"),
)

GOLDEN_0411_SAMPLE: tuple[tuple[str, str, int, str], ...] = (
    ("04:48", "C4", 50000, "카드"),
    ("01:13", "C3", 50000, "카드"),
)


def _fill_sheet(
    ws,
    biz: date,
    sales_rows: Iterable[tuple[str, str, int, str]],
    *,
    revenue_hint: int | None = None,
    count_hint: int | None = None,
) -> None:
    ws["A1"] = "판매일지"
    ws["A2"] = datetime(biz.year, biz.month, biz.day)
    if revenue_hint is not None:
        ws["C2"] = "매출 합계:"
        ws["D2"] = revenue_hint
    if count_hint is not None:
        ws["C3"] = "판매 개수:"
        ws["D3"] = count_hint
    ws["A5"] = "번호"
    ws["B5"] = "모델번호"
    ws["C5"] = "컬러번호"
    ws["D5"] = "금액"
    ws["E5"] = "결제방법"
    for i, (model, color, amt, pay) in enumerate(sales_rows, start=6):
        ws.cell(row=i, column=1, value=i - 5)
        ws.cell(row=i, column=2, value=model)
        ws.cell(row=i, column=3, value=color)
        ws.cell(row=i, column=4, value=amt)
        ws.cell(row=i, column=5, value=pay)


def write_sample_bukchon_xlsx(path: Path) -> None:
    """0410(4건) + 0411(2건) + 0413(헤더만, 판매 없음)."""
    wb = Workbook()
    first = wb.active
    assert first is not None
    wb.remove(first)

    ws410 = wb.create_sheet("0410")
    _fill_sheet(
        ws410,
        date(2026, 4, 10),
        GOLDEN_0410_SALES,
        revenue_hint=200000,
        count_hint=4,
    )

    ws411 = wb.create_sheet("0411")
    _fill_sheet(
        ws411,
        date(2026, 4, 11),
        GOLDEN_0411_SAMPLE,
        revenue_hint=100000,
        count_hint=2,
    )

    ws413 = wb.create_sheet("0413")
    _fill_sheet(ws413, date(2026, 4, 13), (), revenue_hint=0, count_hint=0)

    wb.save(path)

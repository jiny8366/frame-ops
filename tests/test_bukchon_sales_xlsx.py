"""bukchon_sales_xlsx — 북촌 판매일지 → product_code / CSV."""

from __future__ import annotations

from datetime import datetime, time
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook

from lib.bukchon_sales_xlsx import (
    bukchon_sales_xlsx_all_mmdd_to_csv_text,
    bukchon_sales_xlsx_to_csv_text,
    model_and_color_to_product_code,
    normalize_color_code,
    normalize_model_cell_to_text,
)
from lib.sales_import import parse_sales_import_csv, validate_sale_drafts

from tests.bukchon_sample_workbook import GOLDEN_0410_SALES, write_sample_bukchon_xlsx


def test_normalize_color_code() -> None:
    assert normalize_color_code("C2") == "C02"
    assert normalize_color_code("c1") == "C01"
    assert normalize_color_code("C17") == "C17"


def test_normalize_model_cell_excel_time_not_text() -> None:
    """엑셀 시각 서식 → 항상 HH:MM 텍스트(01:01 상품코드 앞부분과 동일)."""
    assert normalize_model_cell_to_text(" 01:01 ") == "01:01"
    assert normalize_model_cell_to_text(time(1, 1)) == "01:01"
    assert normalize_model_cell_to_text(datetime(1899, 12, 30, 1, 1)) == "01:01"
    frac_0101 = (60 + 1) / (24 * 60)
    assert normalize_model_cell_to_text(frac_0101) == "01:01"


def test_model_and_color_to_product_code_time() -> None:
    assert model_and_color_to_product_code(" 01:13", "C2") == "01:13-C02"
    assert model_and_color_to_product_code("7:04", "C2") == "07:04-C02"
    assert model_and_color_to_product_code(time(1, 1), "C1") == "01:01-C01"


def test_xlsx_model_column_stores_excel_time(tmp_path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "0410"
    ws["A1"] = "판매일지"
    ws["A2"] = datetime(2026, 4, 10)
    for col, h in enumerate(["번호", "모델번호", "컬러번호", "금액", "결제방법"], start=1):
        ws.cell(row=5, column=col, value=h)
    ws.cell(row=6, column=1, value=1)
    ws.cell(row=6, column=2, value=time(1, 1))
    ws.cell(row=6, column=3, value="C1")
    ws.cell(row=6, column=4, value=50000)
    ws.cell(row=6, column=5, value="카드")
    path = tmp_path / "time_model.xlsx"
    wb.save(path)
    _biz, csv = bukchon_sales_xlsx_to_csv_text(path, "0410")
    drafts = parse_sales_import_csv(csv)
    assert len(drafts) == 1
    assert drafts[0].lines[0].product_code == "01:01-C01"


def test_model_and_color_to_product_code_cx() -> None:
    assert model_and_color_to_product_code("CX2197", "C17") == "CX2197-C17"


def test_minimal_xlsx_to_csv(tmp_path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "0410"
    ws["A1"] = "판매일지"
    ws["A2"] = datetime(2026, 4, 10, 0, 0, 0)
    ws["A5"] = "번호"
    ws["B5"] = "모델번호"
    ws["C5"] = "컬러번호"
    ws["D5"] = "금액"
    ws["E5"] = "결제방법"
    ws["A6"] = 1
    ws["B6"] = " 01:13"
    ws["C6"] = "C2"
    ws["D6"] = 50000
    ws["E6"] = "현금"
    ws["A7"] = 2
    ws["B7"] = "04:20"
    ws["C7"] = "C2"
    ws["D7"] = 50000
    ws["E7"] = "카드"
    path = tmp_path / "buk.xlsx"
    wb.save(path)

    biz, csv = bukchon_sales_xlsx_to_csv_text(path, "0410")
    assert biz.isoformat() == "2026-04-10"
    drafts = parse_sales_import_csv(csv)
    assert len(drafts) == 2
    assert drafts[0].lines[0].product_code == "01:13-C02"
    assert drafts[0].cash_amount == 50000 and drafts[0].card_amount == 0
    assert drafts[1].card_amount == 50000


def test_golden_0410_four_sales_like_user_file(tmp_path) -> None:
    """Downloads 일지 0410 시트와 동일 4건 패턴."""
    path = tmp_path / "golden.xlsx"
    write_sample_bukchon_xlsx(path)
    biz, csv = bukchon_sales_xlsx_to_csv_text(path, "0410")
    assert biz == datetime(2026, 4, 10).date()
    drafts = parse_sales_import_csv(csv)
    assert len(drafts) == 4
    codes = [d.lines[0].product_code for d in drafts]
    assert codes == ["01:13-C02", "04:20-C02", "07:04-C02", "04:35-C02"]
    pays = [(d.cash_amount, d.card_amount) for d in drafts]
    assert pays == [(50000, 0), (50000, 0), (0, 50000), (0, 50000)]


def test_all_mmdd_merge_skips_empty_0413(tmp_path) -> None:
    path = tmp_path / "multi.xlsx"
    write_sample_bukchon_xlsx(path)
    csv = bukchon_sales_xlsx_all_mmdd_to_csv_text(path)
    drafts = parse_sales_import_csv(csv)
    assert len(drafts) == 4 + 2


def test_xlsx_csv_validate_chain_mock_sb(tmp_path, monkeypatch: pytest.MonkeyPatch) -> None:
    path = tmp_path / "v.xlsx"
    write_sample_bukchon_xlsx(path)
    _, csv = bukchon_sales_xlsx_to_csv_text(path, "0410")
    drafts = parse_sales_import_csv(csv)

    pmap = {
        "01:13-C02": {"id": "p1", "product_code": "01:13-C02", "cost_price": 12000},
        "04:20-C02": {"id": "p2", "product_code": "04:20-C02", "cost_price": 12000},
        "07:04-C02": {"id": "p3", "product_code": "07:04-C02", "cost_price": 12000},
        "04:35-C02": {"id": "p4", "product_code": "04:35-C02", "cost_price": 12000},
    }

    def fake_find(sb, code: str):
        return pmap.get(code)

    monkeypatch.setattr("lib.sales_import.find_product", fake_find)
    monkeypatch.setattr("lib.sales_import.is_business_day_settled", lambda *a, **k: False)

    sb = MagicMock()

    def _chain(data):
        ex = MagicMock()
        ex.data = data
        ch = MagicMock()
        ch.execute.return_value = ex
        ch.limit.return_value = ch
        ch.eq.return_value = ch
        ch.select.return_value = ch
        return ch

    store_chain = _chain([{"id": "sid", "store_code": "BKC01"}])
    empty_idem = _chain([])

    def table(name: str):
        if name == "fo_stores":
            return store_chain
        if name == "fo_sales":
            return empty_idem
        return _chain([])

    sb.table.side_effect = table

    err, warn = validate_sale_drafts(sb, drafts)
    assert err == []
    assert warn == []


def test_golden_tuple_count_matches_constant() -> None:
    assert len(GOLDEN_0410_SALES) == 4
